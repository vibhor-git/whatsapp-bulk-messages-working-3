from flask import Flask, render_template, request, jsonify, redirect, url_for, session, make_response
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS # 1. CORS library import kiya gaya
import io, re, requests, os, csv
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.units import inch
from datetime import datetime

# --- Configuration loaded from Environment Variables (Production Best Practice) ---
WHATSAPP_TOKEN = os.environ.get("WHATSAPP_TOKEN")
WHATSAPP_PHONE_NUMBER_ID = os.environ.get("WHATSAPP_PHONE_NUMBER_ID", "926600003859644")
WEBHOOK_VERIFY_TOKEN = os.environ.get("WEBHOOK_VERIFY_TOKEN", "my_webhook_secret_123")
FLASK_SECRET_KEY = os.environ.get("FLASK_SECRET_KEY", "51e12e6bc2f8fd058fdccd7a83664794")
DATABASE_URL = os.environ.get("DATABASE_URL", "mysql+pymysql://root:vibhor1234@localhost/whatsappdb")

app = Flask(__name__)
app.secret_key = FLASK_SECRET_KEY

# 2. CORS ko apply kiya gaya
CORS(app) 

# Check for Critical Tokens
if not WHATSAPP_TOKEN:
    print("FATAL: WHATSAPP_TOKEN environment variable is not set. API sending will fail.")

# Database Configuration
app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# --- Models ---
class History(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    history_title = db.Column(db.String(255), nullable=False)
    phone_numbers_csv = db.Column(db.Text, nullable=True)
    message_title = db.Column(db.String(255), nullable=False)
    message_body = db.Column(db.Text, nullable=False)
    google_drive_link = db.Column(db.String(500), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    messages = db.relationship('MessageRecord', backref='history', cascade='all, delete-orphan')


class MessageRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    history_id = db.Column(db.Integer, db.ForeignKey('history.id'), nullable=False)
    phone_number = db.Column(db.String(40), nullable=False)
    status = db.Column(db.String(50), default='sent')
    delivered = db.Column(db.Boolean, default=False)
    seen = db.Column(db.Boolean, default=False)
    replied = db.Column(db.Boolean, default=False)
    error_message = db.Column(db.Text, nullable=True)
    sent_at = db.Column(db.DateTime, default=datetime.utcnow)
    whatsapp_message_id = db.Column(db.String(200), nullable=True)


with app.app_context():
    db.create_all()


# -------------------------
# Helpers
# -------------------------
def convert_drive_link(link: str) -> str:
    if not link:
        return link
    match = re.search(r"/d/([A-Za-z0-9_-]+)", link) or re.search(r"id=([A-Za-z0-9_-]+)", link)
    if match:
        return f"https://drive.google.com/uc?export=view&id={match.group(1)}"
    return link


def cell_to_str(value):
    """
    Convert Excel cell value to string digits in a robust way:
    - If value is int -> str(int)
    - If value is float (scientific notation) but whole -> cast to int then str
    - Otherwise str(value)
    """
    if value is None:
        return ''
    # For numbers that excel stores as float (scientific notation)
    if isinstance(value, float):
        # if float is integer-like, convert to int to remove .0 and avoid scientific notation
        if value.is_integer():
            return str(int(value))
        # else convert removing decimal point (phones shouldn't have decimals, but be safe)
        return re.sub(r'\D', '', str(value))
    if isinstance(value, int):
        return str(value)
    # if it's already a string, strip whitespace
    return str(value).strip()


def normalize_phone_raw(s: str):
    """Strip non-digit prefixes and return digits-only or None."""
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return None
    # remove leading + and leading 00, keep digits
    s = re.sub(r'^\+','', s)
    s = re.sub(r'^00','', s)
    cleaned = re.sub(r'\D', '', s)
    if not cleaned:
        return None
    return cleaned


def ensure_country_prefix(number_digits: str, country_code_digits: str):
    """
    REVISED LOGIC (V4): Handles the edge case where the phone number starts with the CC digit.
    Ensures E.164 format (CC + Phone Number) unless the CC is already fully integrated.
    """
    if not number_digits or not country_code_digits:
        # Agar number ya CC mein se koi bhi missing hai, toh number ko jaisa hai waise hi return kar dein.
        return number_digits

    # Agar number pehle hi Country Code se start ho raha hai aur uski length 11 ya 12 se zyada hai, toh woh theek hai.
    if number_digits.startswith(country_code_digits) and len(number_digits) > 10:
        return number_digits

    # Jab CC aur Phone number dono ka pehla digit same ho (jaise 7 & 707...),
    # aur phone number ki length 10 digits ho, toh humein CC forcefully prepend karna hoga
    # kyunki yeh local number hi hai, jismein CC missing hai.
    if len(number_digits) == 10 and number_digits.startswith(country_code_digits):
        # Yeh 10-digit number mein CC missing hai. Prepend CC.
        # Ex: CC='7', Num='7072155666' -> Output: 77072155666
        return country_code_digits + number_digits

    # Baaki sab cases mein: Agar number CC se start nahi hota ya length ka issue hai.
    if not number_digits.startswith(country_code_digits):
        # Agar start hi nahi hota, toh bas prepend kar do (Jaise 51 aur 44 ke liye)
        return country_code_digits + number_digits
    
    # Final Fallback: Waise hi return kar dein
    return number_digits


def send_whatsapp_message(phone, title, body, img_url=None):
    """
    Sends the approved template 'orangetour_christmas' using language en_US.
    If img_url is provided, it will be sent as the header image parameter.
    Returns the parsed JSON on success or {'error': {...}} on failure.
    """
    if not WHATSAPP_TOKEN:
        return {"error": {"message": "Missing WHATSAPP_TOKEN env var"}}

    TEMPLATE_NAME = "orangetour_christmas"
    LANGUAGE_CODE = "en_US"   # match Business Manager: English (US)

    url = f"https://graph.facebook.com/v22.0/{WHATSAPP_PHONE_NUMBER_ID}/messages"
    headers = {
        "Authorization": f"Bearer {WHATSAPP_TOKEN}",
        "Content-Type": "application/json"
    }

    payload = {
        "messaging_product": "whatsapp",
        "to": phone,
        "type": "template",
        "template": {
            "name": TEMPLATE_NAME,
            "language": {"code": LANGUAGE_CODE}
        }
    }

    # If your template expects a header image parameter, include it.
    if img_url:
        payload["template"]["components"] = [
            {
                "type": "header",
                "parameters": [
                    {
                        "type": "image",
                        "image": {"link": img_url}
                    }
                ]
            }
        ]

    try:
        r = requests.post(url, json=payload, headers=headers, timeout=20)
    except Exception as e:
        return {"error": {"message": f"Network error: {e}"}}

    # Always try to parse JSON so we can surface exact API error text
    try:
        resp_json = r.json()
    except Exception:
        return {"error": {"message": f"HTTP {r.status_code} - non-json response", "raw": r.text}}

    # Debug logging to console (useful while developing)
    if r.status_code >= 400 or 'error' in resp_json or resp_json.get('errors'):
        print("WhatsApp API ERROR:", r.status_code, resp_json)   # check your console/logs
        err = resp_json.get('error') or (resp_json.get('errors') and resp_json.get('errors')[0])
        
        # IMPROVED ERROR MESSAGE EXTRACTION
        if err:
            error_message = err.get('message') or str(err)
        else:
            error_message = f"HTTP {r.status_code} - Unknown API error. Raw response: {resp_json}"
            
        return {"error": {"message": error_message}, "raw_response": resp_json}

    # success
    print("WhatsApp API OK:", resp_json)
    return resp_json


# -------------------------
# PDF / Reporting (unchanged)
# -------------------------
def generate_report_pdf(history):
    msgs = MessageRecord.query.filter_by(history_id=history.id).all()
    total_attempted = len(msgs)
    if total_attempted == 0:
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4)
        styles = getSampleStyleSheet()
        doc.build([Paragraph("No messages sent.", styles['Heading1'])])
        buf.seek(0)
        return buf

    delivered = sum(1 for m in msgs if m.delivered)
    seen = sum(1 for m in msgs if m.seen)
    replied = sum(1 for m in msgs if m.replied)
    failed = sum(1 for m in msgs if m.status == 'failed')

    # avoid division by zero
    denom = total_attempted or 1

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=inch)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=20, alignment=1)
    story = [
        Paragraph("WhatsApp Campaign Report", title_style),
        Spacer(1, 12),
        Paragraph("Campaign Details", styles['Heading2']),
        Table([
            ['Title:', history.history_title],
            ['Message:', history.message_title],
            ['Sent At:', history.created_at.strftime("%Y-%m-%d %H:%M")],
            ['Recipients (attempted):', str(total_attempted)]
        ], colWidths=[2*inch, 4.0*inch], style=[('GRID', (0,0), (-1,-1), 1, colors.black)]),
        Spacer(1, 12),
        Paragraph("Delivery Statistics", styles['Heading2']),
        Table([
            ['Metric','Count','%'],
            ['Sent (attempted)', str(total_attempted), f"{(total_attempted/denom)*100:.1f}%"],
            ['Delivered', str(delivered), f"{delivered/denom*100:.1f}%"],
            ['Seen', str(seen), f"{seen/denom*100:.1f}%"],
            ['Replied', str(replied), f"{replied/denom*100:.1f}%"],
            ['Failed', str(failed), f"{failed/denom*100:.1f}%"]
        ], colWidths=[2.0*inch,1.2*inch,1.3*inch], style=[
            ('GRID',(0,0),(-1,-1),1,colors.black),
            ('BACKGROUND',(0,0),(-1,0),colors.darkblue),
            ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke)
        ]),
        Spacer(1, 12),
        Paragraph("Message Body", styles['Heading2']),
        Paragraph(history.message_body, styles['Normal'])
    ]
    doc.build(story)
    buf.seek(0)
    return buf


# -------------------------
# Routes 
# -------------------------
@app.route('/', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        u = request.form.get('username','').strip()
        p = request.form.get('password','').strip()
        if u.lower()=='vibhor' and p=='1234':
            session['logged_in']=True
            return redirect(url_for('index'))
        return render_template('login.html', error="Invalid credentials")
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/main')
def index():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    histories = History.query.order_by(History.id.desc()).all()
    return render_template('index.html', history=histories)


@app.route('/send', methods=['POST'])
def send():
    if not session.get('logged_in'):
        return jsonify(status='error', message='Unauthorized'), 401

    # --- Collect phone numbers from multiple sources: text area CSV, uploaded CSV file, or Excel file ---
    csv_text = request.form.get('phone_numbers_csv','').strip() or ''
    normalized_numbers = []   # list of digit-only strings (phone part)
    per_row_cc = []           # parallel list: country code from Excel row if present (digits) or None
    skipped = []

    # 1) If a CSV file uploaded via input named 'phone_csv' (recommended), parse it (single-column CSV expected)
    if 'phone_csv' in request.files and request.files['phone_csv'].filename:
        try:
            raw = request.files['phone_csv'].read()
            # try decode utf-8, fallback to latin-1
            try:
                text = raw.decode('utf-8')
            except Exception:
                text = raw.decode('latin-1')
            reader = csv.reader(io.StringIO(text))
            for row in reader:
                for cell in row:
                    cell = str(cell).strip()
                    if not cell:
                        continue
                    cleaned = normalize_phone_raw(cell)
                    if cleaned:
                        normalized_numbers.append(cleaned)
                        per_row_cc.append(None)
                    else:
                        skipped.append(cell)
        except Exception as e:
            return jsonify(status='error', message=f'Failed reading uploaded CSV: {e}'), 400

    # 2) Excel file option — here we adapt to your provided format:
    if 'excel_file' in request.files and request.files['excel_file'].filename:
        try:
            wb = load_workbook(io.BytesIO(request.files['excel_file'].read()), data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                # Skip empty rows
                if not any(cell is not None and str(cell).strip() != '' for cell in row):
                    continue
                # If row has at least two columns with values, interpret as [country_code, phone]
                if len(row) >= 2 and row[0] is not None and row[1] is not None:
                    raw_cc = cell_to_str(row[0])
                    raw_phone = cell_to_str(row[1])
                    cc_digits = re.sub(r'\D','', raw_cc) if raw_cc else None
                    phone_digits = normalize_phone_raw(raw_phone)
                    if phone_digits:
                        normalized_numbers.append(phone_digits)
                        per_row_cc.append(cc_digits if cc_digits else None)
                    else:
                        skipped.append(f"{raw_cc},{raw_phone}")
                else:
                    # single-column behaviour: take first non-empty cell as phone
                    first = None
                    for c in row:
                        if c is not None and str(c).strip() != '':
                            first = c
                            break
                    if first is None:
                        continue
                    val = cell_to_str(first)
                    cleaned = normalize_phone_raw(val)
                    if cleaned:
                        normalized_numbers.append(cleaned)
                        per_row_cc.append(None)
                    else:
                        skipped.append(str(first))
        except Exception as e:
            return jsonify(status='error', message=f'Failed reading excel file: {e}'), 400

    # 3) Also parse numbers entered in the text area (old behaviour preserved)
    if csv_text:
        raw_items = re.split(r'[,\n;]+', csv_text)
        for raw in raw_items:
            if not raw:
                continue
            cleaned = normalize_phone_raw(raw)
            if cleaned:
                normalized_numbers.append(cleaned)
                per_row_cc.append(None)
            else:
                skipped.append(raw)

    # Deduplicate while preserving order (we must dedupe both lists)
    seen = set()
    deduped_numbers = []
    deduped_cc = []
    for n, cc in zip(normalized_numbers, per_row_cc):
        key = (n, cc or '')   # treat different cc as different entry
        if key not in seen:
            seen.add(key)
            deduped_numbers.append(n)
            deduped_cc.append(cc)
    normalized_numbers = deduped_numbers
    per_row_cc = deduped_cc

    # Get form fields
    title = request.form.get('message_title','').strip()
    body = request.form.get('message_body','').strip()
    img = convert_drive_link(request.form.get('google_drive_link','').strip())
    htitle = request.form.get('history_title','').strip()
    # default_country_code field ko front-end se lein
    default_cc = request.form.get('default_country_code','').strip()   

    # Basic validation: need at least one number and history title
    if not normalized_numbers:
        return jsonify(status='error', message='No valid phone numbers provided'), 400
    if not htitle:
        return jsonify(status='error', message='History title required'), 400

    # If default_cc provided, normalize to digits; otherwise keep None
    default_cc_digits = re.sub(r'\D','', default_cc) if default_cc else None

    # If some rows do not have cc and default_cc_digits missing, we must require default_cc
    needs_default = any(cc is None for cc in per_row_cc)
    if needs_default and not default_cc_digits:
        return jsonify(status='error', message='Some rows do not include country code. Please provide default_country_code (e.g. 91 for India) in the form.'), 400

    # Build final numbers: use per-row cc if present else default_cc_digits
    final_numbers = []
    for rawnum, row_cc in zip(normalized_numbers, per_row_cc):
        cc_to_use = row_cc if row_cc else default_cc_digits
        
        # Validation for safety
        if not cc_to_use:
            # Agar koi CC available nahi hai, toh number ko jaisa hai waise hi rakhein.
            rec_num = rawnum
        else:
            # Sahi prefixing logic apply karein for E.164 format (CC + Phone Digits)
            rec_num = ensure_country_prefix(rawnum, cc_to_use)
            
        # Basic length check (allow 7-15 digits)
        if rec_num is None or len(rec_num) < 7 or len(rec_num) > 15:
            skipped.append(rawnum)
            continue
            
        final_numbers.append(rec_num)

    if not final_numbers:
        return jsonify(status='error', message='No valid phone numbers after applying country codes'), 400

    # Store normalized numbers in history (so refill works with valid numbers)
    hist = History(history_title=htitle, phone_numbers_csv=",".join(final_numbers),
                    message_title=title or '', message_body=body or '', google_drive_link=img or '')
    db.session.add(hist)
    db.session.commit()   # commit now so MessageRecord can reference hist.id

    res_list = []
    for num in final_numbers:
        # create record first with normalized phone number
        rec = MessageRecord(history_id=hist.id, phone_number=num)
        db.session.add(rec)
        db.session.flush()   # get rec.id if needed

        resp = send_whatsapp_message(num, title, body, img)
        # store raw response in error_message for debugging (even on success)
        try:
            rec.whatsapp_message_id = resp.get('messages',[{}])[0].get('id') if isinstance(resp, dict) else None
        except Exception:
            rec.whatsapp_message_id = None

        if resp.get('error') or not resp.get('messages'):
            rec.status = 'failed'
            # extract message if possible
            err = resp.get('error')
            if isinstance(err, dict):
                # IMPROVED ERROR MESSAGE EXTRACTION
                rec.error_message = err.get('message') or str(err)
            else:
                rec.error_message = str(resp.get('raw_response') or err or 'Unknown error')
            res_list.append(f"{num}: ❌ {rec.error_message}")
        else:
            rec.status = 'sent'
            # store full response for post-mortem (string)
            rec.error_message = str(resp)
            res_list.append(f"{num}: ✅ Sent")
        # no commit here; we will commit after loop for efficiency

    db.session.commit()

    result = {"status": "success", "messages": res_list}
    if skipped:
        result['skipped'] = {"count": len(skipped), "items": skipped}
    return jsonify(result)


@app.route('/report/<int:history_id>')
def report_page(history_id):
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    hist = History.query.get_or_404(history_id)
    msgs = hist.messages or []
    total_attempted = len(msgs)
    total_delivered = sum(1 for m in msgs if m.delivered)
    total_not_delivered = sum(1 for m in msgs if not m.delivered and m.status != 'failed')
    total_seen = sum(1 for m in msgs if m.seen)
    total_not_seen = sum(1 for m in msgs if not m.seen and m.status != 'failed')
    total_replied = sum(1 for m in msgs if m.replied)
    total_failed = sum(1 for m in msgs if m.status == 'failed')

    return render_template('report.html', history=hist,
        total_sent=total_attempted,
        total_delivered=total_delivered,
        total_not_delivered=total_not_delivered,
        total_seen=total_seen,
        total_not_seen=total_not_seen,
        total_replied=total_replied,
        total_failed=total_failed
    )


@app.route('/download-report/<int:history_id>')
def download_report(history_id):
    if not session.get('logged_in'): return redirect(url_for('login'))
    hist = History.query.get_or_404(history_id)
    buf = generate_report_pdf(hist)
    resp = make_response(buf.getvalue())
    resp.headers['Content-Type']='application/pdf'
    fname = "".join(c for c in hist.history_title if c.isalnum() or c==' ').strip()
    resp.headers['Content-Disposition'] = f'attachment; filename="{fname}_report.pdf"'
    return resp


@app.route('/refill/<int:history_id>')
def refill(history_id):
    if not session.get('logged_in'): return redirect(url_for('login'))
    r = History.query.get_or_404(history_id)
    return jsonify(history_title=r.history_title,
                    phone_numbers_csv=r.phone_numbers_csv,
                    message_title=r.message_title,
                    message_body=r.message_body,
                    google_drive_link=r.google_drive_link)


@app.route('/delete/<int:history_id>', methods=['DELETE'])
def delete(history_id):
    if not session.get('logged_in'): return jsonify(status='error',message='Not logged'),401
    rec = History.query.get_or_404(history_id)
    db.session.delete(rec); db.session.commit()
    return jsonify(status='success')


# -------------------------
# Webhook: handle statuses and incoming replies
# -------------------------
@app.route('/webhook', methods=['GET','POST'])
def whatsapp_webhook():
    if request.method == 'GET':
        mode = request.args.get('hub.mode')
        token = request.args.get('hub.verify_token')
        challenge = request.args.get('hub.challenge')
        if mode == 'subscribe' and token == WEBHOOK_VERIFY_TOKEN:
            return challenge, 200
        return 'Verification failed', 403

    data = request.get_json()
    # optional: save webhook payload for debugging
    try:
        with open('last_webhook.json','w') as f:
            import json
            json.dump(data, f, indent=2)
    except Exception:
        pass

    updated = False
    if not data:
        return 'OK', 200

    try:
        for entry in data.get('entry', []):
            for change in entry.get('changes', []):
                value = change.get('value', {})

                # Process statuses (delivery/read/failed)
                for status in value.get('statuses', []):
                    msg_id = status.get('id') or status.get('message_id')
                    recipient = status.get('recipient_id') or status.get('to') or status.get('recipient') or status.get('phone_number')
                    rec = None
                    if msg_id:
                        rec = MessageRecord.query.filter_by(whatsapp_message_id=msg_id).first()
                    if not rec and recipient:
                        recip_norm = normalize_phone_raw(recipient)
                        if recip_norm:
                            suffix = recip_norm[-8:]
                            rec = MessageRecord.query.filter(MessageRecord.phone_number.endswith(suffix)).order_by(MessageRecord.sent_at.desc()).first()

                    if rec:
                        st = status.get('status')
                        # mark booleans and status
                        if st == 'delivered':
                            rec.delivered = True
                            rec.status = 'delivered'
                        elif st in ('read', 'seen'):
                            # Delivered ko bhi true rakhein, kyunki read se pehle delivered hona zaroori hai
                            rec.delivered = True 
                            rec.seen = True
                            rec.status = 'seen'
                        elif st == 'failed':
                            rec.status = 'failed'
                            # try to capture reason
                            reason = status.get('error', {}).get('message') or status.get('reason')
                            rec.error_message = reason or rec.error_message
                        else:
                            # other statuses like 'sent' - store for completeness
                            rec.status = st or rec.status
                        updated = True

                # Process incoming messages (replies). Match by phone number (most robust)
                for message in value.get('messages', []):
                    incoming_from = message.get('from') or message.get('sender') or message.get('wa_id')
                    if not incoming_from:
                        continue
                    incoming_norm = normalize_phone_raw(incoming_from)
                    if not incoming_norm:
                        continue
                    # match by last N digits (8) to be robust against formatting differences
                    suffix = incoming_norm[-8:] if len(incoming_norm) > 8 else incoming_norm
                    rec = MessageRecord.query.filter(MessageRecord.phone_number.endswith(suffix)).order_by(MessageRecord.sent_at.desc()).first()
                    if rec and not rec.replied:
                        rec.replied = True
                        # set status optionally
                        rec.status = 'replied'
                        updated = True

        if updated:
            db.session.commit()

    except Exception as e:
        # keep webhook resilient; don't crash on unexpected payloads
        print(f"Error processing webhook: {e}")

    return 'OK', 200


if __name__=="__main__":
    # Ensure you set the WHATSAPP_TOKEN environment variable 
    # before running this application in production.
    app.run(debug=True)